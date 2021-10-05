from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side


def is_customer_payment(row):
    try:
        return row[4] == "Оплата от заказчиков"
    except AttributeError:
        pass


def is_application_filled(row):
    try:
        return row[26].startswith("Заявка")
    except AttributeError:
        pass


def extract_application_number(value):
    result = value.replace("(для перевозчика) ", "").replace("Заявка ", "")
    return result.split(" ")[0]


def extract_application_date(value):
    result = value.replace("(для перевозчика) ", "").replace("Заявка ", "")
    return result.split(" ")[2]


def main():
    payments_wb = load_workbook("./1.xlsx", read_only=True)
    payments_ws = payments_wb.active

    payment_rows = []
    for row in payments_ws.values:
        if is_customer_payment(row) and is_application_filled(row) and "для перевозчика" not in row[26]:
            payment_rows.append(row)

    apps_numbers = [extract_application_number(num[26]) for num in payment_rows]

    app_wb = load_workbook("./2.xlsx", read_only=True)
    app_ws = app_wb.active

    apps_rows = []
    for row in app_ws.values:
        if row[4].replace(";", "") in apps_numbers:
            apps_rows.append(row)

    result_list = []
    for payment_row in payment_rows:
        for apps_row in apps_rows:
            if apps_row[4].replace(";", "") == extract_application_number(payment_row[26]):
                result_list.append({
                    "customer": apps_row[7],
                    "customer_payment_sum": payment_row[17],
                    "app_number": apps_row[4].replace(";", ""),
                    "app_date": extract_application_date(payment_row[26]),
                    "route": apps_row[5],
                    "transporter": apps_row[10],
                    "transporter_sum": apps_row[8],
                    "nds": apps_row[9]
                })

    result_wb = Workbook()
    result_ws = result_wb.active

    result_ws["A1"] = "Номер заявки"
    result_ws["B1"] = "Дата заявки"
    result_ws["C1"] = "Заказчик"
    result_ws["D1"] = "Ставка заказчика"
    result_ws["E1"] = "Маршрут"
    result_ws["F1"] = "Перевозчик"
    result_ws["G1"] = "Ставка перевозчика"
    result_ws["H1"] = "НДС"
    result_ws["I1"] = "Премия (10%)"
    result_ws["J1"] = "Премия (30%)"

    for record_idx, record in enumerate(result_list, start=2):
        result_ws[f"A{record_idx}"] = record["app_number"]
        result_ws[f"B{record_idx}"] = record["app_date"]
        result_ws[f"C{record_idx}"] = record["customer"]
        result_ws[f"D{record_idx}"] = record["customer_payment_sum"]
        result_ws[f"E{record_idx}"] = record["route"]
        result_ws[f"F{record_idx}"] = record["transporter"]
        result_ws[f"G{record_idx}"] = record["transporter_sum"]

        if "не облагается" in record["nds"]:
            result_ws[f"H{record_idx}"] = "без НДС"
            result_ws[f"I{record_idx}"] = (record["customer_payment_sum"] - record["transporter_sum"]) * 0.1
            result_ws[f"J{record_idx}"] = (record["customer_payment_sum"] / 1.2 - record["transporter_sum"]) * 0.3

    for row_idx, row in enumerate(result_ws, start=2):
        if result_ws[f"i{row_idx}"].value:
            if result_ws[f"I{row_idx}"].value > result_ws[f"J{row_idx}"].value:
                result_ws[f"I{row_idx}"].fill = PatternFill("solid", fgColor="0000FF00")
            else:
                result_ws[f"J{row_idx}"].fill = PatternFill("solid", fgColor="0000FF00")

        for cell in row:
            cell.font = Font("Arial", 8)
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

    for column in result_ws.columns:
        result_ws.column_dimensions[column[0].column_letter].auto_size = True

    result_ws.column_dimensions["C"].width = 25
    result_ws.column_dimensions["D"].width = 14
    result_ws.column_dimensions["E"].width = 40
    result_ws.column_dimensions["F"].width = 40
    result_ws.column_dimensions["G"].width = 16

    result_wb.save("./result.xlsx")


if __name__ == '__main__':
    main()
